"""
Equipment Specs Scraper v4.0
=============================
Busca especificações técnicas de equipamentos comerciais/industriais
(Potência W, Voltagem V, Consumo kWh, BTU) via web scraping.

Equipamentos suportados:
  - Freezers, geladeiras industriais, expositores
  - Fornos industriais/comerciais
  - Ar-condicionado (split, janela, cassete)
  - Coifas industriais
  - Fritadeiras industriais
  - Qualquer equipamento com ficha técnica online

Uso:
    python freezer_specs_scraper.py --txt lista.txt -o resultado.xlsx
"""

import requests
from bs4 import BeautifulSoup
import re
import json
import time
import csv
import random
import argparse
import sys
import os
import tempfile
from typing import Optional
from dataclasses import dataclass, asdict
from urllib.parse import quote_plus, urlparse, urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed

# Força UTF-8 no stdout para evitar erros de encoding no Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    from googlesearch import search as _google_search
    HAS_GOOGLE = True
except ImportError:
    HAS_GOOGLE = False

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        HAS_WEBDRIVER_MANAGER = True
    except ImportError:
        HAS_WEBDRIVER_MANAGER = False
    HAS_SELENIUM = True
except ImportError:
    HAS_SELENIUM = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------

SEARCH_DELAY_MIN = 0.3
SEARCH_DELAY_MAX = 0.8
MAX_RETRIES = 1
REQUEST_TIMEOUT = 8
MAX_URLS_PER_PRODUCT = 10
SELENIUM_WAIT = 4
PDF_MAX_SIZE_MB = 25
PDF_MAX_PAGES = 30

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

EQUIPMENT_CONTEXT_KEYWORDS = [
    "forno", "freezer", "geladeira", "refrigerador", "expositor",
    "ar condicionado", "ar-condicionado", "split", "coifa",
    "fritadeira", "cervejeira", "balcao", "vitrine",
]

QUERY_TEMPLATES = [
    "{product} especificações técnicas potência watts voltagem",
    "{product} ficha técnica",
    "{product} especificações potência consumo",
    "{product} datasheet specifications power watts voltage",
    "{product} manual técnico watts volts",
    "{product} especificaciones técnicas potencia watts voltaje",
    "{product} ficha técnica BTU capacidade",
    "{product} especificações técnicas equipamento comercial",
]

IRRELEVANT_KEYWORDS = [
    "carro", "automóvel", "automovel", "veículo", "veiculo", "sedan",
    "hatch", "suv", "pickup", "motor 1.", "motor 2.", "cilindrada",
    "cambio", "câmbio", "test drive", "concessionária", "concessionaria",
    "ferrari", "lamborghini", "porsche", "maserati", "bugatti",
    "fiat", "chevrolet", "volkswagen", "toyota", "honda motor",
    "bmw", "mercedes-benz", "audi", "hyundai", "nissan",
    "cavalos", "cv de potência", "torque nm", "0 a 100",
    "imóvel", "imovel", "apartamento", "aluguel", "corretor",
    "hotel", "pousada", "hospedagem", "diária", "diaria",
    "passagem", "voo", "aereo", "aéreo",
]

PDF_QUERY_TEMPLATES = [
    "{product} ficha técnica filetype:pdf",
    "{product} catálogo técnico filetype:pdf",
    "{product} datasheet filetype:pdf",
    "{product} manual especificações filetype:pdf",
    "{product} especificações técnicas filetype:pdf",
]

MANUFACTURER_DOMAINS = {
    # ── Refrigeração ──
    "imbera": "https://www.imbera.com/br/produtos?q={q}",
    "metalfrio": "https://www.metalfrio.com.br/busca?q={q}",
    "gelopar": "https://www.gelopar.com.br/busca?search={q}",
    "fricon": "https://www.fricon.com.br/?s={q}",
    "hussmann": "https://www.hussmann.com.br/?s={q}",
    "refrimate": "https://www.refrimate.com.br/busca?q={q}",
    "apolo frio": "https://www.apolofrio.com.br/?s={q}",
    "apolo": "https://www.apolofrio.com.br/?s={q}",
    "mipal": "https://www.mipal.com.br/?s={q}",
    "klima": "https://www.klima.com.br/?s={q}",
    "tecumseh": "https://www.tecumseh.com/br/produtos?q={q}",
    "embraco": "https://www.embraco.com/pt/produtos/?q={q}",
    "imbera mx": "https://www.imbera.com/mx/productos?q={q}",
    # ── Linha branca / multi-categoria ──
    "electrolux": "https://www.electrolux.com.br/search/?q={q}",
    "consul": "https://www.consul.com.br/busca/?q={q}",
    "brastemp": "https://www.brastemp.com.br/busca/?q={q}",
    "midea": "https://www.midea.com.br/busca?q={q}",
    "whirlpool": "https://www.whirlpool.com.br/busca?q={q}",
    "samsung": "https://www.samsung.com.br/busca/?searchvalue={q}",
    "lg": "https://www.lg.com/br/busca/?searchvalue={q}",
    # ── Fornos industriais / cozinha comercial ──
    "tramontina": "https://www.tramontina.com.br/busca?q={q}",
    "braesi": "https://www.braesi.com.br/?s={q}",
    "progás": "https://www.progas.com.br/?s={q}",
    "progas": "https://www.progas.com.br/?s={q}",
    "venâncio": "https://www.venancio.com.br/?s={q}",
    "venancio": "https://www.venancio.com.br/?s={q}",
    "prática": "https://www.praticakitchen.com.br/?s={q}",
    "pratica": "https://www.praticakitchen.com.br/?s={q}",
    "tedesco": "https://www.tedesco.ind.br/?s={q}",
    "rational": "https://www.rational-online.com/br_br/search/?q={q}",
    "metalmaq": "https://www.metalmaq.com.br/?s={q}",
    "gpaniz": "https://www.gpaniz.com.br/?s={q}",
    "skymsen": "https://www.skymsen.com.br/?s={q}",
    # ── Fritadeiras industriais ──
    "croydon": "https://www.croydon.com.br/?s={q}",
    "marchesoni": "https://www.marchesoni.com.br/?s={q}",
    "fritanella": "https://www.fritanella.com.br/?s={q}",
    # ── Coifas ──
    "nardelli": "https://www.nardelli.com.br/?s={q}",
    "suggar": "https://www.suggar.com.br/?s={q}",
    "fischer": "https://www.fischer.com.br/?s={q}",
    "cadence": "https://www.cadence.com.br/busca?q={q}",
    # ── Ar-condicionado ──
    "elgin": "https://www.elgin.com.br/busca?q={q}",
    "carrier": "https://www.carrier.com.br/busca?q={q}",
    "daikin": "https://www.daikin.com.br/pesquisa?q={q}",
    "fujitsu": "https://www.fujitsu-general.com.br/?s={q}",
    "springer": "https://www.springer.com.br/?s={q}",
    "gree": "https://www.gree.com.br/?s={q}",
    "komeco": "https://www.komeco.com.br/?s={q}",
    "agratto": "https://www.agratto.com.br/?s={q}",
    "philco": "https://www.philco.com.br/busca?q={q}",
}

RETAIL_SEARCH_URLS = [
    # Varejo Brasil
    ("Magazine Luiza", "https://www.magazineluiza.com.br/busca/{q}/"),
    ("Zoom", "https://www.zoom.com.br/search?q={q}"),
    ("Buscapé", "https://www.buscape.com.br/search?q={q}"),
    ("Amazon BR", "https://www.amazon.com.br/s?k={q}"),
    ("Mercado Livre", "https://lista.mercadolivre.com.br/{q}"),
    ("Casas Bahia", "https://www.casasbahia.com.br/busca/{q}"),
    ("Ponto", "https://www.pontofrio.com.br/busca/{q}"),
    ("Shopee", "https://shopee.com.br/search?keyword={q}"),
    ("Google Shopping", "https://www.google.com.br/search?q={q}&tbm=shop"),
    ("Carrefour", "https://www.carrefour.com.br/s?q={q}&sort="),
    ("Americanas", "https://www.americanas.com.br/busca/{q}"),
    ("Kabum", "https://www.kabum.com.br/busca/{q}"),
    ("Leroy Merlin", "https://www.leroymerlin.com.br/search?term={q}"),
    ("Submarino", "https://www.submarino.com.br/busca/{q}"),
    ("Girafa", "https://www.girafa.com.br/busca?q={q}"),
    # Especializados em refrigeracao e cozinha comercial
    ("Varimaq", "https://www.varimaq.com.br/buscar?q={q}"),
    ("Refrinox", "https://www.refrinox.com.br/?s={q}"),
    ("Frigelar", "https://www.frigelar.com.br/busca?q={q}"),
    ("ClimaRio", "https://www.climario.com.br/busca?q={q}"),
    ("Frio Peças", "https://www.friopecas.com.br/busca?q={q}"),
    ("Equipamentos Gastronomia", "https://www.emporiogastronomico.com.br/busca?q={q}"),
    ("Solution Inox", "https://www.solutioninox.com.br/busca?q={q}"),
    ("Macom", "https://www.macom.com.br/?s={q}"),
    ("Rede Frio", "https://www.redefrio.com.br/?s={q}"),
    ("WebContinental", "https://www.webcontinental.com.br/busca?q={q}"),
    # Ar-condicionado especializados
    ("WebAr Condicionado", "https://www.webarcondicionado.com.br/busca?q={q}"),
    ("Loja do Mecanico", "https://www.lojadomecanico.com.br/busca?q={q}"),
    # Internacional / Mexico
    ("Mercado Libre MX", "https://listado.mercadolibre.com.mx/{q}"),
    ("Amazon MX", "https://www.amazon.com.mx/s?k={q}"),
    ("Home Depot MX", "https://www.homedepot.com.mx/buscar?query={q}"),
    # Estados Unidos
    ("Best Buy", "https://www.bestbuy.com/site/searchpage.jsp?st={q}"),
    ("Amazon US", "https://www.amazon.com/s?k={q}"),
    ("Home Depot US", "https://www.homedepot.com/s/{q}"),
    ("Lowes", "https://www.lowes.com/search?searchTerm={q}"),
    ("Walmart US", "https://www.walmart.com/search?q={q}"),
]

HP_TO_WATTS = {
    "1/8": 93, "1/6": 124, "1/5": 149, "1/4": 186,
    "1/3": 249, "1/2": 373, "3/4": 559, "1": 746,
    "1.5": 1119, "2": 1492, "3": 2238, "5": 3730,
}


@dataclass
class FreezerSpecs:
    produto: str
    potencia_w: Optional[str] = None
    voltagem_v: Optional[str] = None
    consumo_kwh: Optional[str] = None
    btu: Optional[str] = None
    fase: Optional[str] = None
    fonte_potencia: Optional[str] = None
    fonte_voltagem: Optional[str] = None
    fonte_consumo: Optional[str] = None
    fonte_btu: Optional[str] = None
    fonte_fase: Optional[str] = None


# ---------------------------------------------------------------------------
# HTTP Session & Fetch
# ---------------------------------------------------------------------------

def get_random_headers() -> dict:
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
        "DNT": "1",
    }


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(get_random_headers())
    return session


def fetch_page(url: str, session: Optional[requests.Session] = None,
               retries: int = MAX_RETRIES) -> Optional[str]:
    """Fetch com retry e backoff."""
    s = session or create_session()
    for attempt in range(retries + 1):
        try:
            s.headers["User-Agent"] = random.choice(USER_AGENTS)
            resp = s.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            resp.raise_for_status()
            resp.encoding = resp.apparent_encoding or "utf-8"
            return resp.text
        except requests.exceptions.HTTPError as e:
            if resp.status_code in (403, 429, 503) and attempt < retries:
                wait = (attempt + 1) * 3
                print(f" [retry em {wait}s]", end="")
                time.sleep(wait)
                continue
            return None
        except Exception:
            if attempt < retries:
                time.sleep(2)
                continue
            return None
    return None


_selenium_driver = None

def get_selenium_driver():
    """Inicializa o Selenium WebDriver (headless Chrome)."""
    global _selenium_driver
    if _selenium_driver:
        return _selenium_driver

    if not HAS_SELENIUM:
        return None

    try:
        options = ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        options.add_argument(f"--user-agent={random.choice(USER_AGENTS)}")
        options.add_argument("--lang=pt-BR")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        if HAS_WEBDRIVER_MANAGER:
            service = ChromeService(ChromeDriverManager().install())
            _selenium_driver = webdriver.Chrome(service=service, options=options)
        else:
            _selenium_driver = webdriver.Chrome(options=options)

        _selenium_driver.set_page_load_timeout(20)
        return _selenium_driver
    except Exception as e:
        print(f"\n  [!] Selenium indisponível: {e}")
        return None


def fetch_with_selenium(url: str) -> Optional[str]:
    """Fallback: busca página usando Selenium (renderiza JavaScript)."""
    driver = get_selenium_driver()
    if not driver:
        return None
    try:
        driver.get(url)
        WebDriverWait(driver, SELENIUM_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(2)
        return driver.page_source
    except Exception:
        return None


def close_selenium():
    global _selenium_driver
    if _selenium_driver:
        try:
            _selenium_driver.quit()
        except Exception:
            pass
        _selenium_driver = None


# ---------------------------------------------------------------------------
# PDF — download, parsing e extração de links
# ---------------------------------------------------------------------------

def is_pdf_url(url: str) -> bool:
    """Verifica se a URL aponta para um PDF."""
    parsed = urlparse(url.lower().split("?")[0].split("#")[0])
    return parsed.path.endswith(".pdf")


def download_pdf(url: str, session: Optional[requests.Session] = None) -> Optional[str]:
    """Baixa um PDF para um arquivo temporário. Retorna o caminho ou None."""
    if not HAS_PDF:
        return None
    s = session or create_session()
    try:
        s.headers["User-Agent"] = random.choice(USER_AGENTS)
        resp = s.get(url, timeout=REQUEST_TIMEOUT + 10, stream=True, allow_redirects=True)
        resp.raise_for_status()

        content_type = resp.headers.get("Content-Type", "").lower()
        if "pdf" not in content_type and not is_pdf_url(url):
            return None

        content_length = int(resp.headers.get("Content-Length", 0))
        if content_length > PDF_MAX_SIZE_MB * 1024 * 1024:
            print(f"[PDF muito grande: {content_length // (1024*1024)}MB]", end=" ")
            return None

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        size = 0
        for chunk in resp.iter_content(chunk_size=8192):
            size += len(chunk)
            if size > PDF_MAX_SIZE_MB * 1024 * 1024:
                tmp.close()
                os.unlink(tmp.name)
                return None
            tmp.write(chunk)
        tmp.close()
        return tmp.name
    except Exception:
        return None


def extract_text_from_pdf(pdf_path: str) -> Optional[str]:
    """Extrai texto de um PDF usando pdfplumber."""
    if not HAS_PDF:
        return None
    try:
        import logging
        logging.getLogger("pdfminer").setLevel(logging.ERROR)
        logging.getLogger("pdfplumber").setLevel(logging.ERROR)

        all_text = []
        with pdfplumber.open(pdf_path) as pdf:
            pages_to_read = min(len(pdf.pages), PDF_MAX_PAGES)
            for page in pdf.pages[:pages_to_read]:
                text = page.extract_text()
                if text:
                    all_text.append(text)

                # Também extrai de tabelas dentro do PDF
                tables = page.extract_tables()
                for table in (tables or []):
                    for row in table:
                        if row:
                            row_text = " | ".join(str(cell or "") for cell in row)
                            all_text.append(row_text)

        return "\n".join(all_text) if all_text else None
    except Exception:
        return None
    finally:
        try:
            os.unlink(pdf_path)
        except OSError:
            pass


def extract_specs_from_pdf(url: str,
                           session: Optional[requests.Session] = None) -> dict:
    """Baixa um PDF, extrai texto e busca specs."""
    result = {"potencia": None, "voltagem": None, "consumo": None, "btu": None, "fase": None}
    if not HAS_PDF:
        return result

    pdf_path = download_pdf(url, session)
    if not pdf_path:
        return result

    text = extract_text_from_pdf(pdf_path)
    if not text:
        return result

    result["potencia"] = find_power(text)
    result["voltagem"] = find_voltage(text)
    result["consumo"] = find_consumption(text)
    result["btu"] = find_btu(text)
    result["fase"] = find_phase(text)
    return result


def find_pdf_links_in_page(html: str, base_url: str,
                           product_keywords: Optional[list[str]] = None) -> list[str]:
    """Encontra links para PDFs dentro de uma página HTML."""
    soup = BeautifulSoup(html, "html.parser")
    pdf_links = []
    seen = set()

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        full_url = urljoin(base_url, href)

        if not is_pdf_url(full_url):
            continue
        if full_url in seen:
            continue
        seen.add(full_url)

        # Prioriza PDFs cujo link ou texto contêm palavras-chave
        link_text = (a_tag.get_text(strip=True) + " " + href).lower()
        relevance_keywords = [
            "ficha", "técnica", "tecnica", "catálogo", "catalogo",
            "manual", "especificação", "especificacao", "datasheet",
            "spec", "dados", "download",
        ]
        is_relevant = any(kw in link_text for kw in relevance_keywords)

        if product_keywords:
            has_product = any(kw.lower() in link_text for kw in product_keywords)
            if has_product:
                is_relevant = True

        if is_relevant:
            pdf_links.insert(0, full_url)
        else:
            pdf_links.append(full_url)

    return pdf_links[:10]


def google_search_pdfs(product: str, max_results: int = 5) -> list[str]:
    """Busca Google focada em PDFs (filetype:pdf)."""
    if not HAS_GOOGLE or not HAS_PDF:
        return []

    all_urls = []
    seen = set()
    queries = simplify_product_name(product)

    for variant in queries[:2]:
        for template in PDF_QUERY_TEMPLATES[:2]:
            query = template.format(product=variant)
            try:
                results = list(_google_search(query, num_results=max_results, lang="pt"))
                for url in results:
                    if url not in seen:
                        seen.add(url)
                        all_urls.append(url)
                time.sleep(random.uniform(2, 4))
            except Exception:
                return all_urls

    return all_urls


# ---------------------------------------------------------------------------
# Busca — simplificação de queries
# ---------------------------------------------------------------------------

DIMENSION_PATTERN = re.compile(
    r"\b(?:comp(?:rimento)?|larg(?:ura)?|alt(?:ura)?|prof(?:undidade)?)"
    r"\s*[:\-]?\s*\d+[.,]?\d*\s*(?:m|cm|mm)?\b",
    re.IGNORECASE,
)
DIMENSION_NUMBERS = re.compile(r"\b\d+[.,]\d+\b")
SIZE_PATTERN = re.compile(r"\b\d+\s*[xX]\s*\d+(?:\s*[xX]\s*\d+)?\b")


def _has_equipment_type(product: str) -> bool:
    """Verifica se o nome do produto já contém um tipo de equipamento."""
    lower = product.lower()
    return any(kw in lower for kw in EQUIPMENT_CONTEXT_KEYWORDS)


NEGATIVE_SEARCH_TERMS = ' -carro -ferrari -automóvel -veículo -"test drive"'


def simplify_product_name(product: str) -> list[str]:
    """Gera variações simplificadas do nome do produto para busca."""
    queries = [product]

    simplified = DIMENSION_PATTERN.sub("", product).strip()
    simplified = SIZE_PATTERN.sub("", simplified).strip()
    simplified = DIMENSION_NUMBERS.sub("", simplified).strip()
    simplified = re.sub(r"\s{2,}", " ", simplified).strip()

    if simplified != product and len(simplified) > 5:
        queries.append(simplified)

    words = simplified.split()
    if len(words) > 3:
        short = " ".join(words[:4])
        if short not in queries:
            queries.append(short)

    if _has_equipment_type(product):
        queries = [f"{q} especificações técnicas potência watts{NEGATIVE_SEARCH_TERMS}" for q in queries]
    else:
        queries = [f"{q} equipamento especificações{NEGATIVE_SEARCH_TERMS}" for q in queries]

    return queries


# ---------------------------------------------------------------------------
# Busca — Selenium Google/Bing (fallback robusto)
# ---------------------------------------------------------------------------

def selenium_google_search(query: str, max_results: int = 10) -> list[str]:
    """Usa Selenium para buscar no Google (bypass de bloqueios)."""
    if not HAS_SELENIUM:
        return []
    driver = get_selenium_driver()
    if not driver:
        return []
    try:
        search_url = f"https://www.google.com.br/search?q={quote_plus(query)}&hl=pt-BR"
        driver.get(search_url)
        time.sleep(3)
        soup = BeautifulSoup(driver.page_source, "html.parser")

        urls = []
        seen = set()
        for a_tag in soup.find_all("a", href=True):
            href = a_tag["href"]
            if href.startswith("/url?q="):
                real_url = href.split("/url?q=")[1].split("&")[0]
                from urllib.parse import unquote
                real_url = unquote(real_url)
            elif href.startswith("http") and "google" not in href:
                real_url = href
            else:
                continue

            if any(skip in real_url for skip in
                   ["google.", "youtube.", "gstatic.", "googleapis.", "schema.org"]):
                continue
            if real_url not in seen:
                seen.add(real_url)
                urls.append(real_url)
            if len(urls) >= max_results:
                break
        return urls
    except Exception:
        return []


def _decode_bing_tracking_url(tracking_url: str) -> Optional[str]:
    """Decodifica URL de tracking do Bing para obter a URL real."""
    import base64
    try:
        if "&u=a1" in tracking_url:
            encoded = tracking_url.split("&u=a1")[1].split("&")[0]
            padding = 4 - len(encoded) % 4
            if padding != 4:
                encoded += "=" * padding
            decoded = base64.urlsafe_b64decode(encoded).decode("utf-8", errors="ignore")
            if decoded.startswith("http"):
                return decoded
    except Exception:
        pass
    return None


def selenium_bing_search(query: str, max_results: int = 10) -> list[str]:
    """Usa Selenium para buscar no Bing."""
    if not HAS_SELENIUM:
        return []
    driver = get_selenium_driver()
    if not driver:
        return []
    try:
        import base64
        search_url = f"https://www.bing.com/search?q={quote_plus(query)}&setlang=pt-br&cc=br"
        driver.get(search_url)
        time.sleep(5)

        # Aceita cookies do Bing se houver popup
        try:
            accept_btn = driver.find_element(By.ID, "bnp_btn_accept")
            accept_btn.click()
            time.sleep(2)
        except Exception:
            pass
        try:
            accept_btn = driver.find_element(By.CSS_SELECTOR, "button[id*='accept'], button[id*='consent']")
            accept_btn.click()
            time.sleep(2)
        except Exception:
            pass

        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        urls = []
        seen = set()

        # Estratégia 1: b_algo items com decode de tracking URL
        for li in soup.find_all("li", class_="b_algo"):
            a_tag = li.find("a", href=True)
            if not a_tag:
                continue
            href = a_tag["href"]

            real_url = None
            if "&u=a1" in href:
                try:
                    encoded = href.split("&u=a1")[1].split("&")[0]
                    padding = 4 - len(encoded) % 4
                    if padding != 4:
                        encoded += "=" * padding
                    decoded = base64.urlsafe_b64decode(encoded).decode("utf-8", errors="ignore")
                    if decoded.startswith("http"):
                        real_url = decoded
                except Exception:
                    pass

            if not real_url and href.startswith("http") and "bing.com" not in href:
                real_url = href

            if not real_url:
                cite = li.find("cite")
                if cite:
                    cite_text = cite.get_text(strip=True)
                    if "http" in cite_text:
                        parts = cite_text.split("›")[0].strip().rstrip("/")
                        if parts.startswith("http"):
                            real_url = parts

            if real_url and real_url not in seen:
                seen.add(real_url)
                urls.append(real_url)
            if len(urls) >= max_results:
                break

        # Estratégia 2: fallback — todas as <a> externas na página
        if not urls:
            for a_tag in soup.find_all("a", href=True):
                href = a_tag["href"]
                if "&u=a1" in href:
                    try:
                        encoded = href.split("&u=a1")[1].split("&")[0]
                        padding = 4 - len(encoded) % 4
                        if padding != 4:
                            encoded += "=" * padding
                        decoded = base64.urlsafe_b64decode(encoded).decode("utf-8", errors="ignore")
                        if decoded.startswith("http") and "bing" not in decoded and "microsoft" not in decoded:
                            if decoded not in seen:
                                seen.add(decoded)
                                urls.append(decoded)
                    except Exception:
                        continue
                if len(urls) >= max_results:
                    break

        return urls
    except Exception as e:
        print(f" [Bing erro: {e}]", end="")
        return []


# ---------------------------------------------------------------------------
# Busca — múltiplas estratégias
# ---------------------------------------------------------------------------

def google_search_multi(product: str, max_results: int = 10) -> list[str]:
    """Tenta múltiplos motores de busca em sequência."""
    all_urls = []
    seen = set()
    queries = simplify_product_name(product)

    # 1) Selenium Bing (mais confiável — não tem CAPTCHA)
    if HAS_SELENIUM:
        print("\n    [Bing]", end="", flush=True)
        for variant in queries[:2]:
            query = f"{variant} especificações técnicas potência watts"
            results = selenium_bing_search(query, max_results)
            for url in results:
                if url not in seen:
                    seen.add(url)
                    all_urls.append(url)
            if all_urls:
                break
            time.sleep(2)
        print(f" {len(all_urls)}", end="", flush=True)

    # 2) googlesearch-python (se Bing falhou)
    if not all_urls and HAS_GOOGLE:
        print(" [Google lib]", end="", flush=True)
        for product_variant in queries[:2]:
            for template in QUERY_TEMPLATES[:1]:
                query = template.format(product=product_variant)
                try:
                    results = list(_google_search(query, num_results=max_results, lang="pt"))
                    for url in results:
                        if url not in seen:
                            seen.add(url)
                            all_urls.append(url)
                    time.sleep(random.uniform(2, 4))
                except Exception:
                    break
            if all_urls:
                break
        print(f" {len(all_urls)}", end="", flush=True)

    # 3) Selenium Google (último recurso)
    if not all_urls and HAS_SELENIUM:
        print(" [Google Selenium]", end="", flush=True)
        for variant in queries[:1]:
            query = f"{variant} especificações potência voltagem"
            results = selenium_google_search(query, max_results)
            for url in results:
                if url not in seen:
                    seen.add(url)
                    all_urls.append(url)
            if all_urls:
                break
        print(f" {len(all_urls)}", end="", flush=True)

    return all_urls


def detect_manufacturer(product: str) -> list[str]:
    """Detecta fabricante no nome e retorna URLs do site do fabricante."""
    product_lower = product.lower()
    urls = []
    for brand, url_template in MANUFACTURER_DOMAINS.items():
        if brand in product_lower:
            q = quote_plus(product)
            urls.append(url_template.format(q=q))
            # Também tenta com nome simplificado
            for simplified in simplify_product_name(product)[1:]:
                q2 = quote_plus(simplified)
                candidate = url_template.format(q=q2)
                if candidate not in urls:
                    urls.append(candidate)
    return urls


def build_retail_urls(product: str) -> list[str]:
    """Constrói URLs de busca para sites de varejo com query simplificada."""
    queries = simplify_product_name(product)
    best_query = queries[1] if len(queries) > 1 else queries[0]
    q = quote_plus(best_query)
    return [url.format(q=q) for _, url in RETAIL_SEARCH_URLS]


def get_all_search_urls(product: str) -> tuple[list[str], list[str]]:
    """Combina todas as fontes de URLs. Retorna (html_urls, pdf_urls)."""
    queries = simplify_product_name(product)
    print(f"  Queries: {queries}")

    print("  [1/4] Buscando (Google/Selenium/Bing)...", end="", flush=True)
    google_urls = google_search_multi(product)
    print(f" {len(google_urls)} URLs")

    print("  [2/4] Buscando PDFs...", end="", flush=True)
    pdf_urls = google_search_pdfs(product)
    print(f" {len(pdf_urls)} URLs")

    print("  [3/4] Sites de fabricantes...", end="", flush=True)
    manufacturer_urls = detect_manufacturer(product)
    print(f" {len(manufacturer_urls)} URLs")

    print("  [4/4] Sites de varejo...", end="", flush=True)
    retail_urls = build_retail_urls(product)
    print(f" {len(retail_urls)} URLs")

    all_html = []
    all_pdf = list(pdf_urls)

    for u in google_urls + manufacturer_urls + retail_urls:
        if is_pdf_url(u):
            all_pdf.append(u)
        else:
            all_html.append(u)

    def dedup(lst):
        seen = set()
        out = []
        for u in lst:
            if u not in seen:
                seen.add(u)
                out.append(u)
        return out

    return dedup(all_html)[:MAX_URLS_PER_PRODUCT], dedup(all_pdf)[:10]


# ---------------------------------------------------------------------------
# Extração — JSON-LD
# ---------------------------------------------------------------------------

def extract_from_json_ld(html: str) -> dict:
    """Extrai specs de dados estruturados JSON-LD (schema.org)."""
    result = {"potencia": None, "voltagem": None, "consumo": None, "btu": None, "fase": None}

    soup = BeautifulSoup(html, "html.parser")
    scripts = soup.find_all("script", type="application/ld+json")

    for script in scripts:
        try:
            data = json.loads(script.string)
            if isinstance(data, list):
                for item in data:
                    _parse_json_ld_item(item, result)
            else:
                _parse_json_ld_item(data, result)
        except (json.JSONDecodeError, TypeError):
            continue

    return result


def _parse_json_ld_item(item: dict, result: dict):
    """Processa um item JSON-LD buscando specs."""
    if not isinstance(item, dict):
        return

    props = item.get("additionalProperty", [])
    if not isinstance(props, list):
        props = [props]

    for prop in props:
        if not isinstance(prop, dict):
            continue
        name = str(prop.get("name", "")).lower()
        value = str(prop.get("value", ""))

        if any(k in name for k in ("potência", "potencia", "power", "watt")):
            result["potencia"] = _normalize_power(value)
        elif any(k in name for k in ("tensão", "tensao", "voltagem", "voltage")):
            result["voltagem"] = _normalize_voltage(value)
        elif any(k in name for k in ("consumo", "kwh", "consumption")):
            result["consumo"] = _normalize_consumption(value)
        elif any(k in name for k in ("btu", "capacidade de refrigeração", "cooling capacity")):
            btu_val = find_btu(value)
            if btu_val:
                result["btu"] = btu_val
        elif any(k in name for k in ("fase", "phase", "fases")):
            phase_val = find_phase(value)
            if phase_val:
                result["fase"] = phase_val

    desc = str(item.get("description", ""))
    if desc and len(desc) > 20:
        _extract_from_text_to_result(desc, result)

    if "hasEnergyConsumption" in item:
        energy = item["hasEnergyConsumption"]
        if isinstance(energy, dict):
            val = energy.get("value", "")
            if val:
                result["consumo"] = f"{val} kWh/mês"


# ---------------------------------------------------------------------------
# Extração — Tabelas HTML de especificações
# ---------------------------------------------------------------------------

SPEC_TABLE_KEYWORDS = [
    "especificações", "especificacoes", "ficha técnica", "ficha tecnica",
    "características", "caracteristicas", "specifications", "dados técnicos",
    "informações técnicas", "detalhes técnicos",
]


def extract_from_spec_tables(html: str) -> dict:
    """Busca tabelas de especificações no HTML."""
    result = {"potencia": None, "voltagem": None, "consumo": None, "btu": None, "fase": None}
    soup = BeautifulSoup(html, "html.parser")

    tables = soup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")
        for row in rows:
            cells = row.find_all(["th", "td"])
            if len(cells) >= 2:
                label = cells[0].get_text(strip=True).lower()
                value = cells[1].get_text(strip=True)
                _match_label_value(label, value, result)

    # Muitos sites usam <dl> ao invés de <table>
    dls = soup.find_all("dl")
    for dl in dls:
        dts = dl.find_all("dt")
        dds = dl.find_all("dd")
        for dt, dd in zip(dts, dds):
            label = dt.get_text(strip=True).lower()
            value = dd.get_text(strip=True)
            _match_label_value(label, value, result)

    # Padrão div com label/value (comum em e-commerces)
    for div in soup.find_all("div", class_=re.compile(r"spec|feature|attr|prop|detail|info", re.I)):
        text = div.get_text(separator="|", strip=True)
        parts = text.split("|")
        for i in range(len(parts) - 1):
            label = parts[i].strip().lower()
            value = parts[i + 1].strip()
            _match_label_value(label, value, result)

    return result


def _match_label_value(label: str, value: str, result: dict):
    """Relaciona label+value de uma tabela de specs aos campos desejados."""
    if not value or len(value) > 100:
        return

    if any(k in label for k in ("potência", "potencia", "power", "wattage", "watts")):
        parsed = _normalize_power(value)
        if parsed and not result["potencia"]:
            result["potencia"] = parsed

    elif any(k in label for k in ("tensão", "tensao", "voltagem", "voltage", "tensão nominal")):
        parsed = _normalize_voltage(value)
        if parsed and not result["voltagem"]:
            result["voltagem"] = parsed

    elif any(k in label for k in ("consumo", "kwh", "energia", "consumption", "energy")):
        parsed = _normalize_consumption(value)
        if parsed and not result["consumo"]:
            result["consumo"] = parsed

    elif any(k in label for k in ("btu", "capacidade de refrigeração", "capacidade de refrigeracao",
                                   "cooling capacity", "capacidade frigorífica", "capacidade frigorifica")):
        parsed = find_btu(value)
        if parsed and not result.get("btu"):
            result["btu"] = parsed

    elif any(k in label for k in ("fase", "phase", "fases", "tipo de corrente", "corrente")):
        parsed = find_phase(value)
        if parsed and not result.get("fase"):
            result["fase"] = parsed


# ---------------------------------------------------------------------------
# Extração — Regex em texto livre
# ---------------------------------------------------------------------------

POWER_PATTERNS = [
    r"[Pp]ot[êe]ncia\s*(?:nominal\s*)?(?:\([Ww]\)\s*)?[:\-–\s]*(\d+[\.,]?\d*)\s*[Ww](?:atts?)?",
    r"[Pp]ot[êe]ncia\s+(\d+[\.,]?\d*)\s*[Ww]",
    r"[Pp]ower\s*[:\-–]?\s*(\d+[\.,]?\d*)\s*[Ww]",
    r"[Cc]onsumo\s*(?:de\s+)?[Ee]nergia\s*[:\-–]?\s*(\d+[\.,]?\d*)\s*[Ww]",
    r"[Cc]onsumo\s*[:\-–]\s*(\d+[\.,]?\d*)\s*[Ww]",
    r"[Ww]attage[m]?\s*[:\-–]?\s*(\d+[\.,]?\d*)",
    r"\b(\d{2,4})\s*[Ww]atts?\b",
]

HP_PATTERNS = [
    r"[Pp]ot[êe]ncia\s*(?:do\s+)?(?:compressor\s*)?[:\-–]?\s*(\d+/\d+|\d+[\.,]?\d*)\s*[Hh][Pp]",
    r"(\d+/\d+|\d+[\.,]?\d*)\s*[Hh][Pp]\b",
    r"[Mm]otor\s*(?:de\s+)?(\d+/\d+|\d+[\.,]?\d*)\s*[Hh][Pp]",
]

VOLTAGE_PATTERNS = [
    r"[Tt]ens[ãa]o\s*/?\s*[Vv]oltagem\s*[:\-–]?\s*([\d]+(?:\s*/\s*\d+)?)\s*[Vv]?",
    r"[Vv]oltagem\s*(?:nominal\s*)?[:\-–]?\s*([\d]+(?:\s*/\s*\d+)?)\s*[Vv]?",
    r"[Tt]ens[ãa]o\s*(?:nominal\s*)?(?:\([Vv]\)\s*)?[:\-–]?\s*([\d]+(?:\s*/\s*\d+)?)\s*[Vv]?",
    r"[Vv]oltage\s*[:\-–]?\s*([\d]+(?:\s*/\s*\d+)?)\s*[Vv]?",
    r"\b(1[012][07]\s*/\s*220)\s*[Vv]",
    r"\b(127|110|220|240)\s*[Vv]\b",
]

BIVOLT_PATTERN = re.compile(r"\b[Bb]ivolt(?:e)?\b", re.IGNORECASE)

CONSUMPTION_PATTERNS = [
    r"[Cc]onsumo\s*(?:de\s+)?[Ee]nergia\s*(?:mensal\s*)?[:\-–]?\s*(\d+[\.,]?\d*)\s*[Kk][Ww][Hh](?:/m[êe]s)?",
    r"[Cc]onsumo\s*(?:mensal\s*)?[:\-–]?\s*(\d+[\.,]?\d*)\s*[Kk][Ww][Hh]",
    r"(\d+[\.,]?\d*)\s*[Kk][Ww][Hh]\s*/\s*m[êe]s",
    r"[Ee]nergy\s*[Cc]onsumption\s*[:\-–]?\s*(\d+[\.,]?\d*)\s*[Kk][Ww][Hh]",
]

BTU_PATTERNS = [
    r"[Cc]apacidade\s*(?:de\s+)?[Rr]efrigera[çc][ãa]o\s*[:\-–]?\s*(\d[\d.,]*)\s*BTU",
    r"[Cc]apacidade\s*[:\-–]?\s*(\d[\d.,]*)\s*BTU",
    r"(\d[\d.,]*)\s*BTU(?:[s'']?s?)\s*/?\s*[Hh]",
    r"BTU\s*[:\-–]?\s*(\d[\d.,]*)",
    r"[Cc]ooling\s*[Cc]apacity\s*[:\-–]?\s*(\d[\d.,]*)\s*BTU",
    r"[Pp]ot[êe]ncia\s*(?:frigor[íi]fica\s*)?[:\-–]?\s*(\d[\d.,]*)\s*BTU",
]

PHASE_PATTERNS = [
    (re.compile(r"\b[Tt]rif[áa]sic[oa]\b", re.IGNORECASE), "Trifásico"),
    (re.compile(r"\b3\s*[Ff]ases?\b", re.IGNORECASE), "Trifásico"),
    (re.compile(r"\b380\s*[Vv]\b"), "Trifásico"),
    (re.compile(r"\b[Bb]if[áa]sic[oa]\b", re.IGNORECASE), "Bifásico"),
    (re.compile(r"\b2\s*[Ff]ases?\b", re.IGNORECASE), "Bifásico"),
    (re.compile(r"\b[Mm]onof[áa]sic[oa]\b", re.IGNORECASE), "Monofásico"),
    (re.compile(r"\b1\s*[Ff]ase\b", re.IGNORECASE), "Monofásico"),
]


def _normalize_power(value: str) -> Optional[str]:
    match = re.search(r"(\d+[\.,]?\d*)", value)
    if match:
        num_str = match.group(1).replace(",", ".")
        try:
            num = float(num_str)
            if 30 <= num <= 10000:
                return f"{num_str} W"
        except ValueError:
            pass
    return None


def _normalize_voltage(value: str) -> Optional[str]:
    value_lower = value.lower()
    if "bivolt" in value_lower:
        return "Bivolt (110/220 V)"
    match = re.search(r"(\d{2,3}\s*/\s*\d{2,3})", value)
    if match:
        return f"{match.group(1)} V"
    match = re.search(r"(\d{2,3})", value)
    if match:
        num = int(match.group(1))
        if num in (110, 115, 120, 127, 220, 240):
            return f"{num} V"
    return None


def _normalize_consumption(value: str) -> Optional[str]:
    match = re.search(r"(\d+[\.,]?\d*)", value)
    if match:
        num_str = match.group(1).replace(",", ".")
        return f"{num_str} kWh/mês"
    return None


def _hp_to_watts(hp_str: str) -> Optional[str]:
    """Converte HP para Watts."""
    hp_str = hp_str.strip()
    if hp_str in HP_TO_WATTS:
        watts = HP_TO_WATTS[hp_str]
        return f"{watts} W (≈ {hp_str} HP)"
    try:
        hp_val = float(hp_str.replace(",", "."))
        watts = int(hp_val * 746)
        if 50 <= watts <= 10000:
            return f"{watts} W (≈ {hp_str} HP)"
    except ValueError:
        pass
    return None


def find_power(text: str) -> Optional[str]:
    for pattern in POWER_PATTERNS:
        match = re.search(pattern, text)
        if match:
            value = match.group(1).replace(",", ".")
            try:
                num = float(value)
                if 30 <= num <= 10000:
                    return f"{value} W"
            except ValueError:
                continue

    for pattern in HP_PATTERNS:
        match = re.search(pattern, text)
        if match:
            result = _hp_to_watts(match.group(1))
            if result:
                return result

    return None


def find_voltage(text: str) -> Optional[str]:
    if BIVOLT_PATTERN.search(text):
        return "Bivolt (110/220 V)"
    for pattern in VOLTAGE_PATTERNS:
        match = re.search(pattern, text)
        if match:
            value = match.group(1).strip()
            if "/" in value:
                return f"{value} V"
            try:
                num = int(value)
                if num in (110, 115, 120, 127, 220, 240):
                    return f"{value} V"
            except ValueError:
                continue
    return None


def find_consumption(text: str) -> Optional[str]:
    for pattern in CONSUMPTION_PATTERNS:
        match = re.search(pattern, text)
        if match:
            value = match.group(1).replace(",", ".")
            return f"{value} kWh/mês"
    return None


def find_phase(text: str) -> Optional[str]:
    for pattern, label in PHASE_PATTERNS:
        if pattern.search(text):
            return label
    return None


def find_btu(text: str) -> Optional[str]:
    for pattern in BTU_PATTERNS:
        match = re.search(pattern, text)
        if match:
            raw = match.group(1).replace(".", "").replace(",", "")
            try:
                num = int(raw)
                if 3000 <= num <= 1000000:
                    return f"{num:,} BTU/h".replace(",", ".")
            except ValueError:
                pass
    return None


def _clean_html(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "nav", "header", "footer", "noscript", "svg", "iframe"]):
        tag.decompose()
    return soup.get_text(separator="\n")


def extract_from_text(html: str) -> dict:
    """Extrai specs via regex no texto visível da página."""
    text = _clean_html(html)
    return {
        "potencia": find_power(text),
        "voltagem": find_voltage(text),
        "consumo": find_consumption(text),
        "btu": find_btu(text),
        "fase": find_phase(text),
    }


def _extract_from_text_to_result(text: str, result: dict):
    if not result.get("potencia"):
        result["potencia"] = find_power(text)
    if not result.get("voltagem"):
        result["voltagem"] = find_voltage(text)
    if not result.get("consumo"):
        result["consumo"] = find_consumption(text)
    if not result.get("btu"):
        result["btu"] = find_btu(text)
    if not result.get("fase"):
        result["fase"] = find_phase(text)


# ---------------------------------------------------------------------------
# Relevância — verifica se a página é sobre o produto
# ---------------------------------------------------------------------------

def check_page_relevance(html: str, product_name: str, min_keywords: int = 2) -> bool:
    """Verifica se a página contém palavras-chave suficientes do produto
    e não é sobre assuntos irrelevantes (carros, imóveis, etc.)."""
    text_lower = _clean_html(html).lower()

    irrelevant_count = sum(1 for kw in IRRELEVANT_KEYWORDS if kw in text_lower)
    if irrelevant_count >= 2:
        return False

    keywords = [w.lower() for w in product_name.split() if len(w) > 2]
    if not keywords:
        return True
    matches = sum(1 for kw in keywords if kw in text_lower)

    equipment_signals = [
        "potência", "potencia", "voltagem", "tensão", "tensao",
        "consumo", "watts", "btu", "refrigeração", "refrigeracao",
        "ficha técnica", "ficha tecnica", "especificações", "especificacoes",
    ]
    has_equipment_signal = any(s in text_lower for s in equipment_signals)

    if has_equipment_signal:
        return matches >= 1
    return matches >= min(min_keywords, len(keywords))


# ---------------------------------------------------------------------------
# Extração combinada
# ---------------------------------------------------------------------------

def extract_all_specs(html: str, product_name: Optional[str] = None) -> dict:
    """Executa todas as estratégias de extração e combina os resultados."""
    result = {"potencia": None, "voltagem": None, "consumo": None, "btu": None, "fase": None}

    # Verifica relevância se o nome do produto foi fornecido
    if product_name and not check_page_relevance(html, product_name):
        return result

    # 1) JSON-LD (mais confiável se disponível)
    json_ld = extract_from_json_ld(html)
    for key in result:
        if json_ld.get(key) and not result[key]:
            result[key] = json_ld[key]

    # 2) Tabelas de especificações
    table_specs = extract_from_spec_tables(html)
    for key in result:
        if table_specs.get(key) and not result[key]:
            result[key] = table_specs[key]

    # 3) Regex em texto livre (fallback)
    text_specs = extract_from_text(html)
    for key in result:
        if text_specs.get(key) and not result[key]:
            result[key] = text_specs[key]

    return result


# ---------------------------------------------------------------------------
# Processamento principal
# ---------------------------------------------------------------------------

def _specs_complete(specs: FreezerSpecs) -> bool:
    return bool(specs.potencia_w and specs.voltagem_v and specs.consumo_kwh)


def _apply_result(specs: FreezerSpecs, result: dict, url: str) -> list[str]:
    """Aplica specs encontradas ao objeto e retorna lista do que foi novo."""
    found = []
    if result.get("potencia") and not specs.potencia_w:
        specs.potencia_w = result["potencia"]
        specs.fonte_potencia = url
        found.append(f"Pot={result['potencia']}")
    if result.get("voltagem") and not specs.voltagem_v:
        specs.voltagem_v = result["voltagem"]
        specs.fonte_voltagem = url
        found.append(f"Volt={result['voltagem']}")
    if result.get("consumo") and not specs.consumo_kwh:
        specs.consumo_kwh = result["consumo"]
        specs.fonte_consumo = url
        found.append(f"Consumo={result['consumo']}")
    if result.get("btu") and not specs.btu:
        specs.btu = result["btu"]
        specs.fonte_btu = url
        found.append(f"BTU={result['btu']}")
    if result.get("fase") and not specs.fase:
        specs.fase = result["fase"]
        specs.fonte_fase = url
        found.append(f"Fase={result['fase']}")
    return found


def process_product(product_name: str, session: requests.Session) -> FreezerSpecs:
    """Busca e extrai specs de um produto."""
    print(f"\n{'='*70}")
    print(f"  PRODUTO: {product_name}")
    print(f"{'='*70}")

    specs = FreezerSpecs(produto=product_name)
    html_urls, pdf_urls = get_all_search_urls(product_name)
    product_keywords = [w for w in product_name.split() if len(w) > 2]

    # ── FASE 1: PDFs diretos do Google (mais prováveis de ter ficha técnica) ──
    if pdf_urls and HAS_PDF:
        print(f"\n  FASE 1: PDFs encontrados ({len(pdf_urls)})")
        print(f"  {'─'*60}")
        for i, url in enumerate(pdf_urls, 1):
            domain = urlparse(url).netloc[:35]
            fname = url.split("/")[-1][:25]
            print(f"  [PDF {i}/{len(pdf_urls)}] {domain} / {fname}...", end=" ", flush=True)

            result = extract_specs_from_pdf(url, session)
            found = _apply_result(specs, result, url + " [PDF]")

            if found:
                print(f"✓ {' | '.join(found)}")
            else:
                print("– nada")

            if _specs_complete(specs):
                print(f"\n  >>> Todas as specs encontradas via PDF!")
                _print_product_summary(specs)
                return specs
            _delay()

    # ── FASE 2: Páginas HTML (varejo, fabricantes, Google) ──
    print(f"\n  FASE 2: Páginas HTML ({len(html_urls)})")
    print(f"  {'─'*60}")

    for i, url in enumerate(html_urls, 1):
        domain = urlparse(url).netloc[:45]
        print(f"  [{i:2d}/{len(html_urls)}] {domain}...", end=" ", flush=True)

        html = fetch_page(url, session)
        source = "requests"

        if not html and HAS_SELENIUM:
            html = fetch_with_selenium(url)
            source = "selenium"

        if not html:
            print("✗ falha")
            _delay()
            continue

        # Extrai specs do HTML (com filtro de relevância)
        result = extract_all_specs(html, product_name)
        found = _apply_result(specs, result, url)

        if found:
            print(f"✓ [{source}] {' | '.join(found)}")
        else:
            print(f"– nada [{source}]")

        # Também busca links para PDFs dentro da página
        if HAS_PDF and not _specs_complete(specs):
            page_pdfs = find_pdf_links_in_page(html, url, product_keywords)
            for pdf_url in page_pdfs[:3]:
                pdf_name = pdf_url.split("/")[-1][:30]
                print(f"       └─ PDF: {pdf_name}...", end=" ", flush=True)
                pdf_result = extract_specs_from_pdf(pdf_url, session)
                pdf_found = _apply_result(specs, pdf_result, pdf_url + " [PDF]")
                if pdf_found:
                    print(f"✓ {' | '.join(pdf_found)}")
                else:
                    print("–")
                if _specs_complete(specs):
                    break

        if _specs_complete(specs):
            print(f"\n  >>> Todas as specs encontradas!")
            break

        _delay()

    # ── FASE 3: Selenium em URLs Google (fallback se faltam TODAS as specs) ──
    has_anything = specs.potencia_w or specs.voltagem_v or specs.consumo_kwh
    if not has_anything and not _specs_complete(specs) and HAS_SELENIUM:
        already_visited = {specs.fonte_potencia, specs.fonte_voltagem, specs.fonte_consumo}
        sel_urls = [u for u in html_urls[:4] if u not in already_visited]
        if sel_urls:
            print(f"\n  FASE 3: Selenium em {len(sel_urls)} URLs restantes")
            print(f"  {'─'*60}")
            for url in sel_urls:
                domain = urlparse(url).netloc[:45]
                print(f"  [SEL] {domain}...", end=" ", flush=True)
                html = fetch_with_selenium(url)
                if not html:
                    print("✗")
                    continue
                result = extract_all_specs(html, product_name)
                found = _apply_result(specs, result, url)
                if found:
                    print(f"✓ {' | '.join(found)}")
                else:
                    print("–")

                if HAS_PDF and not _specs_complete(specs):
                    page_pdfs = find_pdf_links_in_page(html, url, product_keywords)
                    for pdf_url in page_pdfs[:2]:
                        pdf_name = pdf_url.split("/")[-1][:30]
                        print(f"       └─ PDF: {pdf_name}...", end=" ", flush=True)
                        pdf_result = extract_specs_from_pdf(pdf_url, session)
                        pdf_found = _apply_result(specs, pdf_result, pdf_url + " [PDF]")
                        if pdf_found:
                            print(f"✓ {' | '.join(pdf_found)}")
                        else:
                            print("–")
                        if _specs_complete(specs):
                            break

                if _specs_complete(specs):
                    break
                _delay()

    _print_product_summary(specs)
    return specs


def _delay():
    time.sleep(random.uniform(SEARCH_DELAY_MIN, SEARCH_DELAY_MAX))


def _print_product_summary(specs: FreezerSpecs):
    print(f"\n  {'─'*60}")
    print(f"  RESUMO: {specs.produto}")
    print(f"    Potência:  {specs.potencia_w or '❌ Não encontrada'}")
    print(f"    Voltagem:  {specs.voltagem_v or '❌ Não encontrada'}")
    print(f"    Fase:      {specs.fase or '–'}")
    print(f"    Consumo:   {specs.consumo_kwh or '❌ Não encontrado'}")
    print(f"    BTU:       {specs.btu or '–'}")
    if specs.fonte_potencia:
        print(f"    Fonte (W): {specs.fonte_potencia[:80]}")
    if specs.fonte_voltagem and specs.fonte_voltagem != specs.fonte_potencia:
        print(f"    Fonte (V): {specs.fonte_voltagem[:80]}")


def process_list(products: list[str]) -> list[FreezerSpecs]:
    session = create_session()
    results = []
    total = len(products)
    for i, product in enumerate(products, 1):
        print(f"\n{'#'*70}")
        print(f"  PROGRESSO: [{i}/{total}]")
        print(f"{'#'*70}")
        result = process_product(product.strip(), session)
        results.append(result)
    close_selenium()
    return results


def search_product(product_name: str) -> dict:
    """API simplificada para uso externo (ex: Slack bot).
    Retorna dict com os resultados da busca."""
    session = create_session()
    try:
        specs = process_product(product_name.strip(), session)
        return asdict(specs)
    finally:
        close_selenium()


def search_products(product_names: list[str]) -> list[dict]:
    """API simplificada para multiplos produtos."""
    results = process_list(product_names)
    return [asdict(r) for r in results]


# ---------------------------------------------------------------------------
# I/O
# ---------------------------------------------------------------------------

COMMON_COLUMNS = [
    "produto", "Produto", "PRODUTO",
    "marca", "Marca", "MARCA",
    "modelo", "Modelo", "MODELO",
    "nome", "Nome", "NOME",
    "descricao", "Descricao", "descrição", "Descrição",
    "product", "Product", "PRODUCT",
    "name", "Name", "NAME",
    "item", "Item", "ITEM",
    "equipamento", "Equipamento", "EQUIPAMENTO",
]


def read_csv_input(filepath: str, column: Optional[str] = None) -> list[str]:
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []

        col = None
        if column and column in headers:
            col = column
        else:
            for pc in COMMON_COLUMNS:
                if pc in headers:
                    col = pc
                    break

        if not col:
            print(f"  [!] Colunas disponíveis no CSV: {headers}")
            print(f"      Use --coluna para especificar qual contém o nome do produto.")
            sys.exit(1)

        print(f"  Usando coluna: '{col}'")
        return [row[col] for row in reader if row.get(col, "").strip()]


def read_txt_input(filepath: str) -> list[str]:
    with open(filepath, "r", encoding="utf-8-sig") as f:
        return [line.strip() for line in f if line.strip()]


def read_excel_input(filepath: str, column: Optional[str] = None) -> list[str]:
    if not HAS_OPENPYXL:
        print("[!] Para ler Excel, instale: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("[!] Planilha vazia.")
        sys.exit(1)

    headers = [str(h or "").strip() for h in rows[0]]

    col_idx = None
    if column:
        for i, h in enumerate(headers):
            if h.lower() == column.lower():
                col_idx = i
                break
    if col_idx is None:
        for i, h in enumerate(headers):
            if h in COMMON_COLUMNS:
                col_idx = i
                break
    if col_idx is None:
        print(f"  [!] Colunas no Excel: {headers}")
        print(f"      Use --coluna para especificar.")
        sys.exit(1)

    print(f"  Usando coluna: '{headers[col_idx]}'")
    products = []
    for row in rows[1:]:
        if col_idx < len(row) and row[col_idx]:
            val = str(row[col_idx]).strip()
            if val:
                products.append(val)
    wb.close()
    return products


def save_results_csv(results: list[FreezerSpecs], output_path: str):
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Produto", "Potência (W)", "Voltagem (V)",
                         "Consumo (kWh/mês)", "Fonte Potência", "Fonte Voltagem"])
        for r in results:
            writer.writerow([
                r.produto,
                r.potencia_w or "",
                r.voltagem_v or "",
                r.consumo_kwh or "",
                r.fonte_potencia or "",
                r.fonte_voltagem or "",
            ])
    print(f"\n  Resultados salvos em: {output_path}")


def save_results_excel(results: list[FreezerSpecs], output_path: str):
    if not HAS_OPENPYXL:
        print("[!] openpyxl não instalado, salvando como CSV.")
        save_results_csv(results, output_path.replace(".xlsx", ".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Specs Freezers"

    headers = ["Produto", "Potência (W)", "Voltagem (V)",
               "Consumo (kWh/mês)", "Fonte Potência", "Fonte Voltagem"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for r in results:
        ws.append([
            r.produto,
            r.potencia_w or "",
            r.voltagem_v or "",
            r.consumo_kwh or "",
            r.fonte_potencia or "",
            r.fonte_voltagem or "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    print(f"\n  Resultados salvos em: {output_path}")


def save_results(results: list[FreezerSpecs], output_path: str):
    if output_path.endswith(".xlsx"):
        save_results_excel(results, output_path)
    else:
        save_results_csv(results, output_path)


def print_results_table(results: list[FreezerSpecs]):
    col_w = 40
    print(f"\n{'='*90}")
    print(f"{'RESULTADOS FINAIS':^90}")
    print(f"{'='*90}")
    print(f"{'Produto':<{col_w}} {'Potência':<18} {'Voltagem':<18} {'Consumo':<14}")
    print(f"{'─'*col_w} {'─'*18} {'─'*18} {'─'*14}")

    for r in results:
        nome = r.produto[:col_w]
        pot = r.potencia_w or "❌ N/A"
        volt = r.voltagem_v or "❌ N/A"
        cons = r.consumo_kwh or "❌ N/A"
        print(f"{nome:<{col_w}} {pot:<18} {volt:<18} {cons:<14}")

    print(f"{'='*90}")

    total = len(results)
    with_power = sum(1 for r in results if r.potencia_w)
    with_voltage = sum(1 for r in results if r.voltagem_v)
    with_consumption = sum(1 for r in results if r.consumo_kwh)
    any_spec = sum(1 for r in results if r.potencia_w or r.voltagem_v or r.consumo_kwh)

    print(f"\n  Estatísticas:")
    print(f"    Potência encontrada:  {with_power}/{total}")
    print(f"    Voltagem encontrada:  {with_voltage}/{total}")
    print(f"    Consumo encontrado:   {with_consumption}/{total}")
    print(f"    Pelo menos 1 spec:    {any_spec}/{total}")


def print_capabilities():
    """Mostra quais bibliotecas estão disponíveis."""
    print("\n  Capacidades do script:")
    print(f"    Google Search:  {'✓' if HAS_GOOGLE else '✗ (pip install googlesearch-python)'}")
    print(f"    Selenium:       {'✓' if HAS_SELENIUM else '✗ (pip install selenium webdriver-manager)'}")
    print(f"    PDF Parser:     {'✓' if HAS_PDF else '✗ (pip install pdfplumber)'}")
    print(f"    Excel (xlsx):   {'✓' if HAS_OPENPYXL else '✗ (pip install openpyxl)'}")
    print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Busca Potência (W), Voltagem (V) e Consumo (kWh) de freezers via web scraping.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python freezer_specs_scraper.py "Freezer Consul CHA31EB 310L"
  python freezer_specs_scraper.py "Electrolux H300" "Metalfrio DA302"
  python freezer_specs_scraper.py --csv freezers.csv -o resultado.csv
  python freezer_specs_scraper.py --csv freezers.csv -o resultado.xlsx
  python freezer_specs_scraper.py --txt lista.txt -o resultado.csv
  python freezer_specs_scraper.py --excel entrada.xlsx -o resultado.xlsx
        """,
    )
    parser.add_argument("produtos", nargs="*", help="Nome(s) do(s) produto(s) para buscar")
    parser.add_argument("--csv", help="CSV de entrada")
    parser.add_argument("--txt", help="TXT de entrada (um produto por linha)")
    parser.add_argument("--excel", help="Excel (.xlsx) de entrada")
    parser.add_argument("--coluna", default=None, help="Coluna com o nome do produto")
    parser.add_argument("--output", "-o", default="freezer_specs_resultado.csv",
                        help="Arquivo de saída (.csv ou .xlsx)")

    args = parser.parse_args()
    products = []

    print_capabilities()

    if args.csv:
        print(f"  Lendo CSV: {args.csv}")
        products = read_csv_input(args.csv, args.coluna)
    elif args.txt:
        print(f"  Lendo TXT: {args.txt}")
        products = read_txt_input(args.txt)
    elif args.excel:
        print(f"  Lendo Excel: {args.excel}")
        products = read_excel_input(args.excel, args.coluna)
    elif args.produtos:
        products = args.produtos
    else:
        parser.print_help()
        print("\n  [!] Forneça produto(s) ou um arquivo de entrada (--csv, --txt, --excel).")
        sys.exit(1)

    print(f"\n  Total de produtos: {len(products)}")
    for i, p in enumerate(products, 1):
        print(f"    {i}. {p}")

    results = process_list(products)
    print_results_table(results)
    save_results(results, args.output)


if __name__ == "__main__":
    main()
